import os
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

def table(dataframe, item, path, filter_column=None, filter_value=None):

    print(f"[PYTHON][table.py] Get file name")
    file_name = os.path.basename(item['table'])

    print(f"[PYTHON][table.py] Save CSV")
    csv_path = os.path.join(path, "CSV")
    os.makedirs(csv_path, exist_ok=True)
    save_path = os.path.join(csv_path, file_name)
    dataframe_str = dataframe.astype(str)
    dataframe_str.to_csv(save_path, index=False, sep=item['delimiter'])

    print(f"[PYTHON][table.py] Format Excel")
    excel_path = os.path.join(path, "Excel")
    os.makedirs(excel_path, exist_ok=True)
    file_base, _ = os.path.splitext(file_name)
    excel_save_path = os.path.join(excel_path, file_base + ".xlsx")
    dataframe.to_excel(excel_save_path, index=False)

    try:

        print(f"[PYTHON][table.py] Load Excel file")
        wb = openpyxl.load_workbook(excel_save_path)
        ws = wb.active

        ws.add_table(
            Table(
                displayName="Table1",
                ref=f"{ws.cell(row=ws.min_row, column=ws.min_column).coordinate}:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}",
                tableStyleInfo=TableStyleInfo(
                    name="TableStyleLight8",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
            )
        )

        print(f"[PYTHON][table.py] Adjust column widths")
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        print(f"[PYTHON][table.py] Save Excel file")
        wb.save(excel_save_path)

    except ImportError:
        print("[PYTHON][table.py] openpyxl is not installed. Excel formatting skipped.")
    except Exception as e:
        print(f"[PYTHON][table.py] Error formatting Excel file: {e}")
    
    print(f"[PYTHON][table.py] Return markdown")
    return dataframe.to_markdown(index=False, tablefmt="pipe", colalign=["left"] * len(dataframe.columns))